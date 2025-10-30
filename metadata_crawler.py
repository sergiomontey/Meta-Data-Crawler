#!/usr/bin/env python3
"""
Multi-Source Metadata Crawler
A desktop application for crawling multiple databases, APIs, and files 
to build a central metadata repository with data dictionary generation 
and lineage mapping.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import sqlite3
from datetime import datetime
import threading
import queue
import os
import sys

# Import required libraries
try:
    import requests
    from sqlalchemy import create_engine, inspect, MetaData
    import pandas as pd
    import networkx as nx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"Missing required library: {e}")
    print("Please install: pip install requests sqlalchemy pandas networkx openpyxl --break-system-packages")
    sys.exit(1)


class MetadataCrawler:
    """Core crawler engine for extracting metadata from multiple sources"""
    
    def __init__(self):
        self.metadata_repo = []
        self.lineage_graph = nx.DiGraph()
        self.data_dictionary = {}
        
    def crawl_database(self, connection_string, db_type='sqlite'):
        """Crawl a database and extract schema metadata"""
        try:
            engine = create_engine(connection_string)
            inspector = inspect(engine)
            
            db_metadata = {
                'source_type': 'database',
                'db_type': db_type,
                'connection': connection_string,
                'tables': [],
                'timestamp': datetime.now().isoformat()
            }
            
            for table_name in inspector.get_table_names():
                table_info = {
                    'name': table_name,
                    'columns': [],
                    'primary_keys': inspector.get_pk_constraint(table_name),
                    'foreign_keys': inspector.get_foreign_keys(table_name),
                    'indexes': inspector.get_indexes(table_name)
                }
                
                for column in inspector.get_columns(table_name):
                    column_info = {
                        'name': column['name'],
                        'type': str(column['type']),
                        'nullable': column['nullable'],
                        'default': column.get('default'),
                        'autoincrement': column.get('autoincrement', False)
                    }
                    table_info['columns'].append(column_info)
                    
                    # Add to data dictionary
                    dict_key = f"{table_name}.{column['name']}"
                    self.data_dictionary[dict_key] = {
                        'table': table_name,
                        'column': column['name'],
                        'data_type': str(column['type']),
                        'nullable': column['nullable'],
                        'source': connection_string
                    }
                
                db_metadata['tables'].append(table_info)
                
                # Add to lineage graph
                self.lineage_graph.add_node(table_name, node_type='table', source='database')
                for fk in table_info['foreign_keys']:
                    ref_table = fk['referred_table']
                    self.lineage_graph.add_edge(table_name, ref_table, 
                                               relationship='foreign_key')
            
            self.metadata_repo.append(db_metadata)
            return True, f"Successfully crawled {len(db_metadata['tables'])} tables"
            
        except Exception as e:
            return False, f"Database crawl error: {str(e)}"
    
    def crawl_api(self, api_url, headers=None):
        """Crawl an API endpoint and extract metadata"""
        try:
            response = requests.get(api_url, headers=headers, timeout=30)
            response.raise_for_status()
            
            data = response.json()
            
            api_metadata = {
                'source_type': 'api',
                'url': api_url,
                'timestamp': datetime.now().isoformat(),
                'status_code': response.status_code,
                'schema': self._infer_schema(data)
            }
            
            self.metadata_repo.append(api_metadata)
            
            # Add to lineage graph
            endpoint_name = api_url.split('/')[-1] or 'root'
            self.lineage_graph.add_node(endpoint_name, node_type='api', source=api_url)
            
            return True, f"Successfully crawled API: {api_url}"
            
        except Exception as e:
            return False, f"API crawl error: {str(e)}"
    
    def crawl_file(self, file_path):
        """Crawl a file and extract metadata"""
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            file_metadata = {
                'source_type': 'file',
                'path': file_path,
                'extension': file_ext,
                'timestamp': datetime.now().isoformat(),
                'size': os.path.getsize(file_path)
            }
            
            if file_ext == '.csv':
                df = pd.read_csv(file_path, nrows=5)
                file_metadata['schema'] = self._infer_dataframe_schema(df)
            elif file_ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, nrows=5)
                file_metadata['schema'] = self._infer_dataframe_schema(df)
            elif file_ext == '.json':
                with open(file_path, 'r') as f:
                    data = json.load(f)
                file_metadata['schema'] = self._infer_schema(data)
            
            self.metadata_repo.append(file_metadata)
            
            # Add to lineage graph
            file_name = os.path.basename(file_path)
            self.lineage_graph.add_node(file_name, node_type='file', source=file_path)
            
            return True, f"Successfully crawled file: {file_name}"
            
        except Exception as e:
            return False, f"File crawl error: {str(e)}"
    
    def _infer_schema(self, data, parent_key=''):
        """Infer schema from JSON data"""
        schema = {}
        
        if isinstance(data, dict):
            for key, value in data.items():
                full_key = f"{parent_key}.{key}" if parent_key else key
                schema[full_key] = {
                    'type': type(value).__name__,
                    'sample': str(value)[:100] if value else None
                }
                if isinstance(value, (dict, list)):
                    schema.update(self._infer_schema(value, full_key))
        elif isinstance(data, list) and data:
            schema[parent_key] = {
                'type': 'list',
                'sample': str(data[0])[:100] if data else None
            }
            if isinstance(data[0], dict):
                schema.update(self._infer_schema(data[0], parent_key))
        
        return schema
    
    def _infer_dataframe_schema(self, df):
        """Infer schema from pandas DataFrame"""
        schema = {}
        for col in df.columns:
            schema[col] = {
                'type': str(df[col].dtype),
                'null_count': int(df[col].isnull().sum()),
                'sample': str(df[col].iloc[0]) if len(df) > 0 else None
            }
        return schema
    
    def generate_data_dictionary(self):
        """Generate comprehensive data dictionary from all crawled sources"""
        full_dict = {}
        
        for item in self.metadata_repo:
            source_type = item['source_type']
            
            if source_type == 'database':
                for table in item.get('tables', []):
                    for col in table['columns']:
                        key = f"{table['name']}.{col['name']}"
                        full_dict[key] = {
                            'source': item.get('connection', 'Unknown'),
                            'source_type': 'database',
                            'table': table['name'],
                            'column': col['name'],
                            'data_type': col['type'],
                            'nullable': col['nullable']
                        }
            
            elif source_type == 'file':
                schema = item.get('schema', {})
                for field, info in schema.items():
                    key = f"{os.path.basename(item['path'])}.{field}"
                    full_dict[key] = {
                        'source': item['path'],
                        'source_type': 'file',
                        'field': field,
                        'data_type': info.get('type', 'Unknown')
                    }
            
            elif source_type == 'api':
                schema = item.get('schema', {})
                for field, info in schema.items():
                    key = f"{item['url']}.{field}"
                    full_dict[key] = {
                        'source': item['url'],
                        'source_type': 'api',
                        'field': field,
                        'data_type': info.get('type', 'Unknown')
                    }
        
        return full_dict
    
    def export_to_excel(self, file_path):
        """Export metadata and data dictionary to Excel"""
        try:
            wb = Workbook()
            
            # Sheet 1: Data Dictionary
            ws1 = wb.active
            ws1.title = "Data Dictionary"
            
            headers = ['Field', 'Source', 'Type', 'Table/File', 'Column', 'Data Type', 'Nullable']
            ws1.append(headers)
            
            # Style headers
            for cell in ws1[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            data_dict = self.generate_data_dictionary()
            for key, value in data_dict.items():
                row = [
                    key,
                    value.get('source', ''),
                    value.get('source_type', ''),
                    value.get('table', value.get('field', '')),
                    value.get('column', ''),
                    value.get('data_type', ''),
                    str(value.get('nullable', ''))
                ]
                ws1.append(row)
            
            # Auto-adjust column widths
            for column in ws1.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws1.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Metadata Summary
            ws2 = wb.create_sheet("Metadata Summary")
            ws2.append(['Source Type', 'Count', 'Details'])
            
            # Style headers
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            summary = {}
            for item in self.metadata_repo:
                source_type = item['source_type']
                summary[source_type] = summary.get(source_type, 0) + 1
            
            for source_type, count in summary.items():
                ws2.append([source_type, count, f"{count} {source_type}(s) crawled"])
            
            # Sheet 3: Lineage Information
            ws3 = wb.create_sheet("Lineage Map")
            ws3.append(['From', 'To', 'Relationship'])
            
            for cell in ws3[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for edge in self.lineage_graph.edges(data=True):
                ws3.append([edge[0], edge[1], edge[2].get('relationship', 'related')])
            
            wb.save(file_path)
            return True, f"Successfully exported to {file_path}"
            
        except Exception as e:
            return False, f"Export error: {str(e)}"
    
    def get_lineage_visualization_data(self):
        """Get lineage data for visualization"""
        nodes = []
        edges = []
        
        for node, data in self.lineage_graph.nodes(data=True):
            nodes.append({
                'id': node,
                'type': data.get('node_type', 'unknown'),
                'source': data.get('source', '')
            })
        
        for edge in self.lineage_graph.edges(data=True):
            edges.append({
                'from': edge[0],
                'to': edge[1],
                'relationship': edge[2].get('relationship', 'related')
            })
        
        return nodes, edges


class MetadataCrawlerGUI:
    """Main GUI application for the Metadata Crawler"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Multi-Source Metadata Crawler")
        self.root.geometry("1200x800")
        
        self.crawler = MetadataCrawler()
        self.result_queue = queue.Queue()
        
        self._setup_ui()
        self._check_queue()
        
    def _setup_ui(self):
        """Setup the user interface"""
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Main container
        main_container = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Left panel - Input sources
        left_frame = ttk.Frame(main_container)
        main_container.add(left_frame, weight=1)
        
        # Right panel - Results and visualization
        right_frame = ttk.Frame(main_container)
        main_container.add(right_frame, weight=2)
        
        # Left panel content
        self._setup_input_panel(left_frame)
        
        # Right panel content
        self._setup_results_panel(right_frame)
        
    def _setup_input_panel(self, parent):
        """Setup the input panel for adding sources"""
        # Title
        title_label = ttk.Label(parent, text="Data Sources", font=('Arial', 14, 'bold'))
        title_label.pack(pady=10)
        
        # Notebook for different source types
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Database tab
        db_frame = ttk.Frame(notebook)
        notebook.add(db_frame, text="Database")
        self._setup_database_tab(db_frame)
        
        # API tab
        api_frame = ttk.Frame(notebook)
        notebook.add(api_frame, text="API")
        self._setup_api_tab(api_frame)
        
        # File tab
        file_frame = ttk.Frame(notebook)
        notebook.add(file_frame, text="Files")
        self._setup_file_tab(file_frame)
        
        # Action buttons
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=5, pady=10)
        
        ttk.Button(button_frame, text="Generate Data Dictionary", 
                  command=self.generate_dictionary).pack(fill=tk.X, pady=2)
        ttk.Button(button_frame, text="View Lineage Map", 
                  command=self.view_lineage).pack(fill=tk.X, pady=2)
        ttk.Button(button_frame, text="Export to Excel", 
                  command=self.export_excel).pack(fill=tk.X, pady=2)
        ttk.Button(button_frame, text="Clear All Data", 
                  command=self.clear_data).pack(fill=tk.X, pady=2)
        
    def _setup_database_tab(self, parent):
        """Setup database connection tab"""
        # Connection type
        ttk.Label(parent, text="Database Type:").pack(anchor=tk.W, padx=10, pady=(10,0))
        self.db_type_var = tk.StringVar(value='sqlite')
        db_types = ['sqlite', 'postgresql', 'mysql', 'mssql', 'oracle']
        ttk.Combobox(parent, textvariable=self.db_type_var, values=db_types, 
                    state='readonly').pack(fill=tk.X, padx=10, pady=5)
        
        # Connection string
        ttk.Label(parent, text="Connection String:").pack(anchor=tk.W, padx=10, pady=(10,0))
        self.db_conn_var = tk.StringVar()
        ttk.Entry(parent, textvariable=self.db_conn_var).pack(fill=tk.X, padx=10, pady=5)
        
        # Example text
        example_text = ("Examples:\n"
                       "SQLite: sqlite:///path/to/database.db\n"
                       "PostgreSQL: postgresql://user:pass@localhost/dbname\n"
                       "MySQL: mysql://user:pass@localhost/dbname")
        example_label = ttk.Label(parent, text=example_text, font=('Arial', 8), 
                                 foreground='gray', justify=tk.LEFT)
        example_label.pack(anchor=tk.W, padx=10, pady=5)
        
        # Browse button for SQLite
        ttk.Button(parent, text="Browse SQLite File", 
                  command=self.browse_sqlite).pack(fill=tk.X, padx=10, pady=5)
        
        # Crawl button
        ttk.Button(parent, text="Crawl Database", 
                  command=self.crawl_database, 
                  style='Accent.TButton').pack(fill=tk.X, padx=10, pady=20)
        
    def _setup_api_tab(self, parent):
        """Setup API endpoint tab"""
        # API URL
        ttk.Label(parent, text="API Endpoint URL:").pack(anchor=tk.W, padx=10, pady=(10,0))
        self.api_url_var = tk.StringVar()
        ttk.Entry(parent, textvariable=self.api_url_var).pack(fill=tk.X, padx=10, pady=5)
        
        # Headers (optional)
        ttk.Label(parent, text="Headers (JSON format, optional):").pack(anchor=tk.W, padx=10, pady=(10,0))
        self.api_headers_text = scrolledtext.ScrolledText(parent, height=6)
        self.api_headers_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.api_headers_text.insert(1.0, '{\n  "Authorization": "Bearer YOUR_TOKEN"\n}')
        
        # Example
        example_text = "Example: https://api.example.com/data"
        ttk.Label(parent, text=example_text, font=('Arial', 8), 
                 foreground='gray').pack(anchor=tk.W, padx=10, pady=5)
        
        # Crawl button
        ttk.Button(parent, text="Crawl API", 
                  command=self.crawl_api,
                  style='Accent.TButton').pack(fill=tk.X, padx=10, pady=20)
        
    def _setup_file_tab(self, parent):
        """Setup file selection tab"""
        ttk.Label(parent, text="Select files to crawl:").pack(anchor=tk.W, padx=10, pady=10)
        
        # File list
        self.file_listbox = tk.Listbox(parent, selectmode=tk.MULTIPLE)
        self.file_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Buttons
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Button(btn_frame, text="Add Files", 
                  command=self.add_files).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Remove Selected", 
                  command=self.remove_files).pack(side=tk.LEFT, padx=2)
        
        # Supported formats
        formats_text = "Supported: CSV, Excel (.xlsx, .xls), JSON"
        ttk.Label(parent, text=formats_text, font=('Arial', 8), 
                 foreground='gray').pack(anchor=tk.W, padx=10, pady=5)
        
        # Crawl button
        ttk.Button(parent, text="Crawl Files", 
                  command=self.crawl_files,
                  style='Accent.TButton').pack(fill=tk.X, padx=10, pady=20)
        
    def _setup_results_panel(self, parent):
        """Setup results and visualization panel"""
        # Title
        title_label = ttk.Label(parent, text="Results & Metadata", font=('Arial', 14, 'bold'))
        title_label.pack(pady=10)
        
        # Notebook for different views
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Log tab
        log_frame = ttk.Frame(notebook)
        notebook.add(log_frame, text="Activity Log")
        self._setup_log_tab(log_frame)
        
        # Data Dictionary tab
        dict_frame = ttk.Frame(notebook)
        notebook.add(dict_frame, text="Data Dictionary")
        self._setup_dictionary_tab(dict_frame)
        
        # Lineage tab
        lineage_frame = ttk.Frame(notebook)
        notebook.add(lineage_frame, text="Lineage Map")
        self._setup_lineage_tab(lineage_frame)
        
        # Statistics tab
        stats_frame = ttk.Frame(notebook)
        notebook.add(stats_frame, text="Statistics")
        self._setup_stats_tab(stats_frame)
        
    def _setup_log_tab(self, parent):
        """Setup activity log tab"""
        self.log_text = scrolledtext.ScrolledText(parent, wrap=tk.WORD, 
                                                   font=('Courier', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Add welcome message
        self.log("Multi-Source Metadata Crawler initialized")
        self.log("Ready to crawl databases, APIs, and files")
        self.log("-" * 60)
        
    def _setup_dictionary_tab(self, parent):
        """Setup data dictionary view"""
        # Treeview for dictionary
        columns = ('Field', 'Source', 'Type', 'Data Type', 'Nullable')
        self.dict_tree = ttk.Treeview(parent, columns=columns, show='tree headings')
        
        for col in columns:
            self.dict_tree.heading(col, text=col)
            self.dict_tree.column(col, width=150)
        
        # Scrollbars
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.dict_tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self.dict_tree.xview)
        self.dict_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.dict_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        
    def _setup_lineage_tab(self, parent):
        """Setup lineage visualization tab"""
        self.lineage_text = scrolledtext.ScrolledText(parent, wrap=tk.WORD, 
                                                      font=('Courier', 9))
        self.lineage_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
    def _setup_stats_tab(self, parent):
        """Setup statistics tab"""
        self.stats_text = scrolledtext.ScrolledText(parent, wrap=tk.WORD, 
                                                    font=('Courier', 10))
        self.stats_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
    def log(self, message):
        """Add message to activity log"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.update_idletasks()
        
    def browse_sqlite(self):
        """Browse for SQLite database file"""
        filename = filedialog.askopenfilename(
            title="Select SQLite Database",
            filetypes=[("SQLite Database", "*.db *.sqlite *.sqlite3"), ("All Files", "*.*")]
        )
        if filename:
            conn_string = f"sqlite:///{filename}"
            self.db_conn_var.set(conn_string)
            self.log(f"Selected database: {filename}")
            
    def add_files(self):
        """Add files to crawl list"""
        filenames = filedialog.askopenfilenames(
            title="Select Files to Crawl",
            filetypes=[
                ("Supported Files", "*.csv *.xlsx *.xls *.json"),
                ("CSV Files", "*.csv"),
                ("Excel Files", "*.xlsx *.xls"),
                ("JSON Files", "*.json"),
                ("All Files", "*.*")
            ]
        )
        for filename in filenames:
            if filename not in self.file_listbox.get(0, tk.END):
                self.file_listbox.insert(tk.END, filename)
                self.log(f"Added file: {os.path.basename(filename)}")
                
    def remove_files(self):
        """Remove selected files from list"""
        selected = self.file_listbox.curselection()
        for index in reversed(selected):
            filename = self.file_listbox.get(index)
            self.file_listbox.delete(index)
            self.log(f"Removed file: {os.path.basename(filename)}")
            
    def crawl_database(self):
        """Crawl database in background thread"""
        conn_string = self.db_conn_var.get().strip()
        if not conn_string:
            messagebox.showwarning("Input Required", "Please enter a connection string")
            return
        
        db_type = self.db_type_var.get()
        self.log(f"Starting database crawl: {db_type}")
        
        def _crawl():
            success, message = self.crawler.crawl_database(conn_string, db_type)
            self.result_queue.put(('database', success, message))
        
        thread = threading.Thread(target=_crawl, daemon=True)
        thread.start()
        
    def crawl_api(self):
        """Crawl API in background thread"""
        api_url = self.api_url_var.get().strip()
        if not api_url:
            messagebox.showwarning("Input Required", "Please enter an API URL")
            return
        
        # Parse headers
        headers = None
        headers_text = self.api_headers_text.get(1.0, tk.END).strip()
        if headers_text:
            try:
                headers = json.loads(headers_text)
            except json.JSONDecodeError:
                messagebox.showerror("Invalid JSON", "Headers must be valid JSON")
                return
        
        self.log(f"Starting API crawl: {api_url}")
        
        def _crawl():
            success, message = self.crawler.crawl_api(api_url, headers)
            self.result_queue.put(('api', success, message))
        
        thread = threading.Thread(target=_crawl, daemon=True)
        thread.start()
        
    def crawl_files(self):
        """Crawl files in background thread"""
        files = list(self.file_listbox.get(0, tk.END))
        if not files:
            messagebox.showwarning("No Files", "Please add files to crawl")
            return
        
        self.log(f"Starting file crawl: {len(files)} file(s)")
        
        def _crawl():
            results = []
            for file_path in files:
                success, message = self.crawler.crawl_file(file_path)
                results.append((success, message))
            self.result_queue.put(('files', results))
        
        thread = threading.Thread(target=_crawl, daemon=True)
        thread.start()
        
    def generate_dictionary(self):
        """Generate and display data dictionary"""
        if not self.crawler.metadata_repo:
            messagebox.showinfo("No Data", "Please crawl some sources first")
            return
        
        self.log("Generating data dictionary...")
        
        # Clear existing items
        for item in self.dict_tree.get_children():
            self.dict_tree.delete(item)
        
        # Generate dictionary
        data_dict = self.crawler.generate_data_dictionary()
        
        # Populate tree
        for key, value in data_dict.items():
            self.dict_tree.insert('', tk.END, text=key, values=(
                key,
                value.get('source', '')[:50],
                value.get('source_type', ''),
                value.get('data_type', ''),
                str(value.get('nullable', ''))
            ))
        
        self.log(f"Data dictionary generated: {len(data_dict)} entries")
        messagebox.showinfo("Success", f"Generated data dictionary with {len(data_dict)} entries")
        
    def view_lineage(self):
        """Display lineage map"""
        if not self.crawler.lineage_graph.nodes():
            messagebox.showinfo("No Data", "No lineage information available")
            return
        
        self.log("Generating lineage map...")
        
        # Clear existing text
        self.lineage_text.delete(1.0, tk.END)
        
        # Get lineage data
        nodes, edges = self.crawler.get_lineage_visualization_data()
        
        # Display nodes
        self.lineage_text.insert(tk.END, "=== DATA ENTITIES ===\n\n")
        for node in nodes:
            self.lineage_text.insert(tk.END, 
                f"• {node['id']} ({node['type']})\n  Source: {node['source']}\n\n")
        
        # Display relationships
        self.lineage_text.insert(tk.END, "\n=== RELATIONSHIPS ===\n\n")
        for edge in edges:
            self.lineage_text.insert(tk.END, 
                f"{edge['from']} → {edge['to']} ({edge['relationship']})\n")
        
        self.log(f"Lineage map displayed: {len(nodes)} entities, {len(edges)} relationships")
        
    def export_excel(self):
        """Export to Excel file"""
        if not self.crawler.metadata_repo:
            messagebox.showinfo("No Data", "Please crawl some sources first")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Export to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if filename:
            self.log(f"Exporting to Excel: {filename}")
            success, message = self.crawler.export_to_excel(filename)
            
            if success:
                self.log("Export completed successfully")
                messagebox.showinfo("Success", message)
            else:
                self.log(f"Export failed: {message}")
                messagebox.showerror("Export Error", message)
                
    def clear_data(self):
        """Clear all crawled data"""
        if messagebox.askyesno("Confirm", "Clear all crawled data?"):
            self.crawler = MetadataCrawler()
            
            # Clear UI elements
            for item in self.dict_tree.get_children():
                self.dict_tree.delete(item)
            self.lineage_text.delete(1.0, tk.END)
            self.stats_text.delete(1.0, tk.END)
            
            self.log("All data cleared")
            self.update_statistics()
            
    def update_statistics(self):
        """Update statistics display"""
        self.stats_text.delete(1.0, tk.END)
        
        self.stats_text.insert(tk.END, "=== METADATA CRAWLER STATISTICS ===\n\n")
        
        # Count sources
        source_counts = {}
        for item in self.crawler.metadata_repo:
            source_type = item['source_type']
            source_counts[source_type] = source_counts.get(source_type, 0) + 1
        
        self.stats_text.insert(tk.END, "Sources Crawled:\n")
        for source_type, count in source_counts.items():
            self.stats_text.insert(tk.END, f"  • {source_type.capitalize()}: {count}\n")
        
        # Data dictionary size
        data_dict = self.crawler.generate_data_dictionary()
        self.stats_text.insert(tk.END, f"\nData Dictionary Entries: {len(data_dict)}\n")
        
        # Lineage information
        nodes = len(self.crawler.lineage_graph.nodes())
        edges = len(self.crawler.lineage_graph.edges())
        self.stats_text.insert(tk.END, f"\nLineage Entities: {nodes}\n")
        self.stats_text.insert(tk.END, f"Lineage Relationships: {edges}\n")
        
    def _check_queue(self):
        """Check for results from background threads"""
        try:
            while True:
                result = self.result_queue.get_nowait()
                
                if result[0] == 'database':
                    _, success, message = result
                    if success:
                        self.log(f"✓ {message}")
                    else:
                        self.log(f"✗ {message}")
                        messagebox.showerror("Database Error", message)
                    self.update_statistics()
                    
                elif result[0] == 'api':
                    _, success, message = result
                    if success:
                        self.log(f"✓ {message}")
                    else:
                        self.log(f"✗ {message}")
                        messagebox.showerror("API Error", message)
                    self.update_statistics()
                    
                elif result[0] == 'files':
                    _, results = result
                    for success, message in results:
                        if success:
                            self.log(f"✓ {message}")
                        else:
                            self.log(f"✗ {message}")
                    self.update_statistics()
                    
        except queue.Empty:
            pass
        
        # Schedule next check
        self.root.after(100, self._check_queue)


def main():
    """Main entry point"""
    root = tk.Tk()
    app = MetadataCrawlerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
